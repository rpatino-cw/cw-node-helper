"""IB connection trace — parse xlsx, cache, and search InfiniBand connections."""
from __future__ import annotations

import json
import os
import re
from datetime import datetime

__all__ = [
    '_load_connections', '_search_connections', '_auto_detect_type',
]

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

_XLSX_PATH = os.path.expanduser(
    "~/dev/Network_guides/DH1 & DH2 All_IB_Connections_Simplified_v2.xlsx"
)
_CACHE_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    ".ibtrace_cache.json",
)

# Type prefix map
_TYPE_PREFIX = {"Spine": "S", "Core": "C", "Leaf": "L", "Node": "N"}


# ---------------------------------------------------------------------------
# Normalization helpers
# ---------------------------------------------------------------------------

def _normalize_port(value) -> str:
    """Convert datetime-corrupted port values back to 'port/lane' format.

    Excel m/d number format causes openpyxl to read some port numbers as
    datetime objects where month=port, day=lane.
    """
    if isinstance(value, datetime):
        return f"{value.month}/{value.day}"
    if isinstance(value, (int, float)):
        i = int(value)
        return str(i) if value == i else str(value)
    if value is None:
        return ""
    return str(value).strip()


def _normalize_id(value) -> str:
    """Normalize switch ID — floats like 1.1 stay '1.1', strings pass through."""
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:g}"
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _normalize_cab(value) -> str:
    """Cabinet number — always a string."""
    if value is None:
        return ""
    if isinstance(value, float):
        i = int(value)
        return str(i) if value == i else str(value)
    return str(value).strip()


def _build_switch_name(
    type_str: str, switch_id: str, data_hall: str, cab: str = "",
) -> str:
    """Build canonical switch name like 'S1.1.1', 'L10.1.2-DH2', 'C1.4'.

    Leaf names in the burndown format prepend the cabinet number:
      cab=10, id=1.2, DH2 → L10.1.2-DH2
    Spines/Cores just use the ID: S8.3.2, C1.4
    """
    prefix = _TYPE_PREFIX.get(type_str, "")
    if type_str == "Leaf" and cab:
        name = f"{prefix}{cab}.{switch_id}"
        if data_hall:
            name += f"-{data_hall}"
    else:
        name = f"{prefix}{switch_id}"
    return name


# ---------------------------------------------------------------------------
# XLSX parsing
# ---------------------------------------------------------------------------

def _parse_xlsx(path: str) -> list[dict]:
    """Parse the IB Connections sheet into a list of connection dicts."""
    try:
        import openpyxl
    except ImportError:
        print("  openpyxl is required: pip install openpyxl")
        return []

    if not os.path.isfile(path):
        print(f"  IB connections file not found: {path}")
        return []

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    connections: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 11:
            continue
        # Skip empty rows
        if row[1] is None and row[4] is None:
            continue

        src_type = str(row[1]).strip() if row[1] else ""
        src_dh = str(row[2]).strip() if row[2] else ""
        src_id = _normalize_id(row[4])
        src_port = _normalize_port(row[5])
        dest_type = str(row[6]).strip() if row[6] else ""
        dest_dh = str(row[7]).strip() if row[7] else ""
        dest_id = _normalize_id(row[9])
        dest_port = _normalize_port(row[10])
        tab_ref = str(row[11]).strip() if len(row) > 11 and row[11] else ""

        src_cab = _normalize_cab(row[3])
        dest_cab = _normalize_cab(row[8])

        conn = {
            "data_hall": str(row[0]).strip() if row[0] else "",
            "src_type": src_type,
            "src_dh": src_dh,
            "src_cab": src_cab,
            "src_id": src_id,
            "src_port": src_port,
            "src_name": _build_switch_name(src_type, src_id, src_dh, src_cab),
            "dest_type": dest_type,
            "dest_dh": dest_dh,
            "dest_cab": dest_cab,
            "dest_id": dest_id,
            "dest_port": dest_port,
            "dest_name": _build_switch_name(dest_type, dest_id, dest_dh, dest_cab),
            "tab_ref": tab_ref,
        }
        connections.append(conn)

    wb.close()
    return connections


# ---------------------------------------------------------------------------
# Caching (mtime-based, same pattern as ib_burndown.py)
# ---------------------------------------------------------------------------

def _load_cache(xlsx_path: str) -> list[dict] | None:
    """Return cached connections if cache exists and xlsx hasn't changed."""
    if not os.path.isfile(_CACHE_PATH):
        return None
    try:
        with open(_CACHE_PATH, "r") as f:
            data = json.load(f)
        cached_mtime = data.get("xlsx_mtime", 0)
        current_mtime = os.path.getmtime(xlsx_path)
        if cached_mtime == current_mtime:
            return data.get("connections", [])
    except (json.JSONDecodeError, OSError, KeyError):
        pass
    return None


def _save_cache(connections: list[dict], xlsx_path: str) -> None:
    """Write connections + xlsx mtime to cache file."""
    try:
        data = {
            "xlsx_mtime": os.path.getmtime(xlsx_path),
            "connections": connections,
        }
        with open(_CACHE_PATH, "w") as f:
            json.dump(data, f)
    except OSError:
        pass  # silently skip if can't write


def _load_connections(xlsx_path: str | None = None) -> list[dict]:
    """Load connections from cache or parse xlsx. Returns list of dicts."""
    path = xlsx_path or _XLSX_PATH

    cached = _load_cache(path)
    if cached is not None:
        return cached

    connections = _parse_xlsx(path)
    if connections:
        _save_cache(connections, path)
    return connections


# ---------------------------------------------------------------------------
# Search
# ---------------------------------------------------------------------------

def _switch_matches(name: str, pattern: str) -> bool:
    """Check if a switch name matches a search pattern.

    Handles DH-suffix flexibility: 'L10.1.2' matches 'L10.1.2-DH1' and 'L10.1.2-DH2'.
    """
    name_up = name.upper()
    pat_up = pattern.upper()

    if name_up == pat_up:
        return True
    # Pattern without DH suffix matches name with DH suffix
    if "-" in name_up and "-" not in pat_up:
        base = name_up.split("-")[0]
        return base == pat_up
    return False


def _auto_detect_type(raw_id: str) -> list[str]:
    """Given a bare ID like '8.3.2', return candidate switch names with prefixes.

    3-level (x.y.z) → Spine
    2-level (x.y) → Leaf, then Core
    """
    parts = raw_id.split(".")
    if len(parts) == 3:
        return [f"S{raw_id}"]
    elif len(parts) == 2:
        return [f"L{raw_id}", f"C{raw_id}"]
    elif len(parts) == 1:
        return [f"C{raw_id}"]
    return [raw_id]


def _search_connections(
    connections: list[dict],
    switch: str,
    port: str | None = None,
) -> list[dict]:
    """Find connections where switch matches src or dest side.

    Args:
        connections: full connection list from _load_connections
        switch: switch name (S8.3.2, L10.1.2-DH2, 8.3.2)
        port: optional port filter (22/1)

    Returns:
        list of matching connection dicts
    """
    switch = switch.strip()

    # Determine if input has a type prefix
    has_prefix = bool(re.match(r'^[SCLNscln]', switch)) and not switch[0].isdigit()

    if has_prefix:
        candidates = [switch.upper()]
    else:
        candidates = [c.upper() for c in _auto_detect_type(switch)]

    port_upper = port.upper().strip() if port else None

    results = []
    for conn in connections:
        src_name = conn["src_name"].upper()
        dest_name = conn["dest_name"].upper()

        matched_side = None
        for cand in candidates:
            if _switch_matches(src_name, cand):
                matched_side = "src"
                break
            if _switch_matches(dest_name, cand):
                matched_side = "dest"
                break

        if matched_side is None:
            continue

        # Port filter
        if port_upper:
            if matched_side == "src":
                if conn["src_port"].upper() != port_upper:
                    continue
            else:
                if conn["dest_port"].upper() != port_upper:
                    continue

        results.append(conn)

    return results
