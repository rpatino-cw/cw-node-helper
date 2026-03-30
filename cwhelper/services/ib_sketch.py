"""IB Sketch elevation data — parse ELEV sheets from the IB Sketch xlsx."""
from __future__ import annotations

import json
import os

__all__ = ['_load_ib_sketch', '_get_rack_switches']

_XLSX_PATHS = [
    os.path.expanduser("~/dev/cw-node-helper/source/IB Sketch.xlsx"),
    os.path.expanduser("~/dev/Network_guides/IB Sketch.xlsx"),
]
_CACHE_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    ".ib_sketch_cache.json",
)


def _parse_elev_sheets(xlsx_path: str) -> dict:
    """Parse ELEV sheets → {rack_num: {ru: {name, model}, ...}, ...}

    Also stores row metadata per rack for display headers.
    Returns {"racks": {...}, "switch_to_rack": {...}}
    """
    try:
        import openpyxl
    except ImportError:
        return {}

    if not os.path.isfile(xlsx_path):
        return {}

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    racks: dict[str, dict] = {}  # "121" → {"ru": {38: {"name": "S1.4.1", "model": "MQM9790-NS2F"}, ...}, "row_label": "DH2 Row 13", "config": "IB x16a"}
    switch_to_rack: dict[str, dict] = {}  # "S1.4.1" → {"rack": 121, "ru": 38, "model": "..."}

    for sheet_name in wb.sheetnames:
        if 'ELEV' not in sheet_name and not sheet_name.startswith('DH3 Row'):
            continue

        ws = wb[sheet_name]
        rows_list = list(ws.iter_rows(min_row=1, max_row=55, values_only=True))
        if len(rows_list) < 4:
            continue

        # Row 2: has DH/row/config info (e.g., "IB x16a", "Row 13")
        row2 = rows_list[1] if len(rows_list) > 1 else ()
        config_label = ""
        row_label = sheet_name  # fallback
        if row2 and len(row2) >= 3:
            parts = [str(v).strip() for v in row2[:4] if v is not None]
            config_label = " ".join(parts[1:]) if len(parts) > 1 else ""
            row_label = sheet_name.replace(" ELEV IB x16a", "").strip()

        # Row 3: rack numbers in columns
        header = list(rows_list[2])
        rack_cols: dict[int, int] = {}
        for ci, val in enumerate(header):
            if isinstance(val, (int, float)) and val >= 10:
                rack_cols[ci] = int(val)

        if not rack_cols:
            continue

        for ci, rack_num in rack_cols.items():
            rack_key = str(rack_num)
            if rack_key not in racks:
                racks[rack_key] = {
                    "ru": {},
                    "row_label": row_label,
                    "config": config_label,
                }

            for row_data in rows_list[3:]:
                if not row_data or row_data[0] is None:
                    continue
                ru = int(row_data[0])
                if ci < len(row_data) and row_data[ci] is not None:
                    sw_name = str(row_data[ci]).strip()
                    if not sw_name or sw_name == 'None':
                        continue
                    model = ""
                    if ci + 1 < len(row_data) and row_data[ci + 1]:
                        model = str(row_data[ci + 1]).strip()
                    racks[rack_key]["ru"][str(ru)] = {
                        "name": sw_name,
                        "model": model,
                    }
                    switch_to_rack[sw_name.upper()] = {
                        "rack": rack_num,
                        "ru": ru,
                        "model": model,
                    }

    wb.close()
    return {"racks": racks, "switch_to_rack": switch_to_rack}


def _load_cache(xlsx_path: str) -> dict | None:
    if not os.path.isfile(_CACHE_PATH):
        return None
    try:
        with open(_CACHE_PATH, "r") as f:
            data = json.load(f)
        if data.get("xlsx_mtime") == os.path.getmtime(xlsx_path):
            return data
    except (json.JSONDecodeError, OSError, KeyError):
        pass
    return None


def _save_cache(data: dict, xlsx_path: str) -> None:
    try:
        data["xlsx_mtime"] = os.path.getmtime(xlsx_path)
        with open(_CACHE_PATH, "w") as f:
            json.dump(data, f)
    except OSError:
        pass


def _load_ib_sketch() -> dict:
    """Load IB sketch elevation data (cached). Returns {"racks": {...}, "switch_to_rack": {...}}."""
    for path in _XLSX_PATHS:
        if not os.path.isfile(path):
            continue
        cached = _load_cache(path)
        if cached is not None:
            return cached
        data = _parse_elev_sheets(path)
        if data:
            _save_cache(data, path)
            return data
    return {"racks": {}, "switch_to_rack": {}}


def _get_rack_switches(rack_num: int | str) -> dict | None:
    """Get all switches in a rack from IB sketch. Returns {"ru": {...}, "row_label": "...", "config": "..."} or None."""
    data = _load_ib_sketch()
    return data.get("racks", {}).get(str(rack_num))


def _find_switch_location(switch_name: str) -> dict | None:
    """Find a switch's rack/RU location. Returns {"rack": int, "ru": int, "model": str} or None."""
    data = _load_ib_sketch()
    return data.get("switch_to_rack", {}).get(switch_name.upper())
