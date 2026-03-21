"""RMA tracker — reads the EVI01 Device Tracker xlsx from ~/Downloads.

Workflow: Download the sheet → walkthrough auto-finds it in Downloads → reads it → deletes it.
No GCP, no auth, no API keys, no manual file moving.
"""
from __future__ import annotations

import glob
import os
import re
from datetime import datetime

__all__ = ['_rma_available', '_get_rma_data', '_rma_file_age', '_rma_file_age_secs']

_DOWNLOADS = os.path.expanduser("~/Downloads")
_NODE_RE = re.compile(r'(DH\d+)-R(\d+)-(.*)', re.IGNORECASE)

# Column indices in the "Active Devices" sheet (0-based, row 3 = headers)
_COL = {
    "location":       0,  # DH1-R317-Node-07
    "serial":         1,  # S948338X5616385
    "dh":             2,  # DH1
    "issue":          3,  # Uncabled for RMA.
    "date_reported":  4,  # datetime
    "days_pending":   5,  # formula or number
    "rma_ticket":     6,  # HO-89952
    "status":         7,  # RMA Engaged
    "last_noted":     8,  # datetime
    "assigned_to":    9,
    "notes":         10,
}


# ── Public API ───────────────────────────────────────────────────────────────

def _rma_available() -> bool:
    """True if the tracker xlsx exists in Downloads."""
    return bool(_find_latest_file())


def _get_rma_data(dh: str) -> dict[str, list[dict]]:
    """Find newest tracker in Downloads, read it, filter by DH, clean up.

    Returns {rack_label: [items]}.
    """
    file_path = _find_latest_file()
    if not file_path:
        return {}

    if file_path.lower().endswith(".csv"):
        rows = _load_csv(file_path)
    else:
        rows = _load_xlsx(file_path)

    if not rows:
        return {}

    # Normalize DH input: "Data Hall 1" → "DH1", "dh1" → "DH1", etc.
    dh_normalized = _normalize_dh(dh)
    result: dict[str, list[dict]] = {}

    for row in rows:
        parsed = _parse_node_name(row.get("node_name", ""))
        if not parsed:
            row_dh = _normalize_dh(row.get("dh") or "")
            if row_dh != dh_normalized:
                continue
            continue

        if _normalize_dh(parsed["dh"]) != dh_normalized:
            continue

        rack_label = f"R{int(parsed['rack']):03d}"
        result.setdefault(rack_label, []).append(row)

    return result


def _rma_file_age() -> str:
    """Human-readable age of the newest tracker file."""
    file_path = _find_latest_file()
    if not file_path:
        return "no file"
    try:
        import time
        age_secs = time.time() - os.path.getmtime(file_path)
        if age_secs < 3600:
            return f"{int(age_secs // 60)}m ago"
        elif age_secs < 86400:
            return f"{int(age_secs // 3600)}h ago"
        else:
            return f"{int(age_secs // 86400)}d ago"
    except OSError:
        return "unknown"


def _rma_file_age_secs() -> float:
    """Age of the newest tracker file in seconds. Returns -1 if no file."""
    file_path = _find_latest_file()
    if not file_path:
        return -1
    try:
        import time
        return time.time() - os.path.getmtime(file_path)
    except OSError:
        return -1



# ── DH Normalization ─────────────────────────────────────────────────────────

def _normalize_dh(dh: str) -> str:
    """Normalize data hall names: 'Data Hall 1' → 'DH1', 'dh1' → 'DH1', etc."""
    s = dh.strip()
    # "Data Hall 1" / "Data Hall 2" → "DH1" / "DH2"
    m = re.match(r'data\s*hall\s*(\d+)', s, re.IGNORECASE)
    if m:
        return f"DH{m.group(1)}"
    # Already short form: "DH1", "dh2", "DH-1"
    m = re.match(r'dh[- _]?(\d+)', s, re.IGNORECASE)
    if m:
        return f"DH{m.group(1)}"
    return s.upper()


# ── File Discovery ───────────────────────────────────────────────────────────

def _find_latest_file() -> str | None:
    """Find the newest Device Tracker file (xlsx or csv) in ~/Downloads.

    Handles duplicates: "name (1).xlsx", "name (2).csv", etc.
    """
    patterns = [
        "EVI01-Device-Tracker*.xlsx",
        "EVI01-Device-Tracker*.csv",
        "EVI01*Device*Tracker*.xlsx",
        "EVI01*Device*Tracker*.csv",
        "*Device-Tracker*.xlsx",
        "*Device-Tracker*.csv",
        "*Device*Tracker*.xlsx",
        "*Device*Tracker*.csv",
        "*Active Devices*.csv",
    ]

    candidates = set()
    for pattern in patterns:
        for path in glob.glob(os.path.join(_DOWNLOADS, pattern)):
            candidates.add(path)

    if not candidates:
        return None

    # Return the most recently modified file
    return max(candidates, key=lambda p: os.path.getmtime(p))



# ── CSV Parsing ──────────────────────────────────────────────────────────────

def _load_csv(path: str) -> list[dict]:
    """Read the Active Devices CSV export."""
    import csv
    import io

    try:
        with open(path, newline='', encoding='utf-8-sig') as f:
            text = f.read()
    except OSError:
        return []

    reader = csv.reader(io.StringIO(text))
    all_rows = list(reader)

    # Find the header row (contains "Location" in first column)
    header_idx = None
    for i, row in enumerate(all_rows):
        if row and row[0].strip().lower() == "location":
            header_idx = i
            break

    if header_idx is None:
        return []

    rows_out = []
    for raw_row in all_rows[header_idx + 1:]:
        if not raw_row or not raw_row[0].strip():
            continue
        location = raw_row[0].strip()
        if not location:
            continue

        row = {
            "node_name":     location,
            "serial":        _safe_col(raw_row, 1),
            "dh":            _safe_col(raw_row, 2),
            "issue":         _safe_col(raw_row, 3),
            "date_reported": _safe_col(raw_row, 4),
            "age_days":      _safe_col(raw_row, 5),
            "ho_ticket":     _safe_col(raw_row, 6),
            "status":        _safe_col(raw_row, 7),
            "last_updated":  _safe_col(raw_row, 8),
            "assigned_to":   _safe_col(raw_row, 9),
            "notes":         _safe_col(raw_row, 10),
        }
        rows_out.append(row)

    return rows_out


def _safe_col(row: list, idx: int) -> str:
    """Safely get a column value as string."""
    if idx < len(row):
        return (row[idx] or "").strip()
    return ""


# ── XLSX Parsing ─────────────────────────────────────────────────────────────

def _load_xlsx(path: str) -> list[dict]:
    """Read the Active Devices sheet from the tracker xlsx."""
    try:
        import openpyxl
    except ImportError:
        return []

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return []

    # Find the Active Devices sheet
    sheet_name = None
    for name in wb.sheetnames:
        if "active" in name.lower():
            sheet_name = name
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[0]  # fallback to first sheet

    ws = wb[sheet_name]
    rows_out = []

    # Find the header row (contains "Location" in column A)
    header_row_idx = None
    all_rows = list(ws.iter_rows(values_only=True))
    for i, row in enumerate(all_rows):
        if row and row[0] and str(row[0]).strip().lower() == "location":
            header_row_idx = i
            break

    wb.close()

    if header_row_idx is None:
        return []

    # Parse data rows after header
    for raw_row in all_rows[header_row_idx + 1:]:
        if not raw_row or not raw_row[0]:
            continue
        location = str(raw_row[0] or "").strip()
        if not location or location == " ":
            continue

        def _xcol(col_key: str) -> str:
            idx = _COL[col_key]
            if idx < len(raw_row) and raw_row[idx] is not None:
                return str(raw_row[idx]).strip()
            return ""

        row = {
            "node_name":     location,
            "serial":        _xcol("serial"),
            "dh":            _xcol("dh"),
            "issue":         _xcol("issue"),
            "date_reported": _fmt_date(raw_row[_COL["date_reported"]]) if _COL["date_reported"] < len(raw_row) else "",
            "age_days":      _fmt_age(raw_row[_COL["days_pending"]]) if _COL["days_pending"] < len(raw_row) else "",
            "ho_ticket":     _xcol("rma_ticket"),
            "status":        _xcol("status"),
            "last_updated":  _fmt_date(raw_row[_COL["last_noted"]]) if _COL["last_noted"] < len(raw_row) else "",
            "assigned_to":   _xcol("assigned_to"),
            "notes":         _xcol("notes"),
        }

        rows_out.append(row)

    return rows_out


def _fmt_date(val) -> str:
    """Format a cell value as a date string."""
    if isinstance(val, datetime):
        return val.strftime("%m/%d/%Y")
    if val:
        return str(val).strip()
    return ""


def _fmt_age(val) -> str:
    """Format days pending — could be a number, formula string, or None."""
    if val is None:
        return ""
    if isinstance(val, (int, float)):
        return str(int(val))
    s = str(val).strip()
    # If it's a formula like "=TODAY()-E6", we can't evaluate it — skip
    if s.startswith("="):
        return ""
    return s


def _parse_node_name(name: str) -> dict | None:
    """Parse 'DH1-R317-Node-07' into {dh, rack, suffix}."""
    m = _NODE_RE.match(name.strip())
    if not m:
        return None
    return {
        "dh":     m.group(1).upper(),
        "rack":   m.group(2),
        "suffix": m.group(3),
    }
